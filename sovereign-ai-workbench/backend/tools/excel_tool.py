import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from backend.config import config

class ExcelTool:
    """
    Spreadsheet Agent Tool.
    Reads, calculates, aggregates, and outputs structured .xlsx workbooks using openpyxl.
    """

    def process_and_generate(self, headers: list, rows: list, filename: str = "Plant_Maintenance_Analysis.xlsx") -> str:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Maintenance Analysis"

        # Styling
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        cell_font = Font(name="Arial", size=10)
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        # Write Headers
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(len(str(h)) + 4, 18)

        # Write Rows
        for row_idx, r in enumerate(rows, start=2):
            for col_idx, val in enumerate(r, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = cell_font
                cell.border = thin_border
                if isinstance(val, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")

        output_path = os.path.join(config.DELIVERABLES_DIR, filename)
        wb.save(output_path)
        return output_path

excel_tool = ExcelTool()
